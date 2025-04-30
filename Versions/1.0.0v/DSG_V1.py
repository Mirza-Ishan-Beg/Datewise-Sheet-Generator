from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Border, Side


class DSG:
    def __init__(self, filename, year):
        print("""
        
        \t\t\t+----------------------------------------------------------------+
        \t\t\t|                  CREATED BY AT DATE 07-DEC-2024                |
        \t\t\t+----------------------------------------------------------------+
        \t\t\t|                                                                |
        \t\t\t|          |\\    /|   |   +----+    ------     +-----+           |
        \t\t\t|          | \\  / |   |   |    |        /      |     |           |
        \t\t\t|          |  \\/  |   |   |----+       /       +-----+           |
        \t\t\t|          |      |   |   |  \\        /        |     |           |
        \t\t\t|          |      |   |   |   \\      /_____    |     |           |
        \t\t\t+----------------------------------------------------------------+
        
        
            """)
        self.filename = filename
        self.year = year
        self.days_per_month_2sqr = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        self.days_per_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
        dt_tuple = (year, 1, 1, 0, 0, 0)

        self.monthly_date_chart = [
            "JAN", "FEB", "MAR", "APR",
            "MAY", "JUN", "JUL", "AUG",
            "SEP", "OCT", "NOV", "DEC"
        ]
        self.pure_init_date = datetime(*dt_tuple)
        self.dict_per_blueprint = {}
        self.data_frames = {
            "JAN": "", "FEB": "", "MAR": "", "APR": "",
            "MAY": "", "JUN": "", "JUL": "", "AUG": "",
            "SEP": "", "OCT": "", "NOV": "", "DEC": ""
        }

    def component1(self):
        print("*\t\t\tCOMPONENT 1 WORKING...")
        if year % 4 == 0:
            for i in range(len(self.days_per_month_2sqr)):
                # print(" ===== ===== NEXT MONTH ===== =====")
                dt_tuple = (year, i + 1, 1, 0, 0, 0)
                init_date = datetime(*dt_tuple)
                self.dict_per_blueprint = {
                    str(init_date): [
                        "TOTAL SALES", "CASH SALE", "CREDIT CARD SALES",
                        "PYTM SALE", "LOYALTY REDEEM", "RETURN SALES",
                        "HOLD BILLS", "UNHOLD BILLS", "EXPENSES",
                        "DEEPAWALI PURCHASE", "DRY FRUIT PURCHASE", "STORE RENT",
                        "SUB TOTALS"
                    ],
                    "SUMMARY": [
                        0,
                        "=OFFSET(B$2,(ROW()-3)/13*13,0,1,1)-SUM(OFFSET(B$4,(ROW()-3)/13*13,0,10,1))+OFFSET(B$9,(ROW()-3)/13*13,0,1,1)*2",
                        0, 0, 0, 0,
                        "=SUM(OFFSET(F$2, (ROW() - 8)/12*12, 0, 12, 1))",
                        "=SUM(OFFSET(H$2, (ROW() - 9)/12*12, 0, 12, 1))",
                        "=SUM(OFFSET(J$2, (ROW() - 10)/12*12, 0, 12, 1), OFFSET(L$2, (ROW() - 10)/13*13, 0, 13, 1))",
                        0, 0, 0, ""
                    ],
                    "CASH RECEIVED": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "BANK DEPOSIT": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "NAMES OF HOLD BILLS": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        "SUB TOTAL HOLD"
                    ],
                    "AMOUNT HOLD": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        "=SUM(M2:M31)"
                    ],
                    "NAMES OF UNHOLD BILLS": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        "SUB TOTAL UNHOLD"
                    ],
                    "AMOUNT UNHOLD": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        "=SUM(N2:N31)"
                    ],
                    "CASH USED UP BY STORE CONVENIENCES": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        ""
                    ],
                    "AMOUNT USED 1": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "CASH USED UP IN PURCHASE": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        ""
                    ],
                    "AMOUNT USED 2": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "BILL HOLDING COLUMN": [],
                    "BILL UNHOLDING COLUMN": [],
                    "TOTALS COLUMN": []
                }

                for days in range(2, self.days_per_month_2sqr[i] + 1):
                    dt_tuple = (year, i + 1, days, 0, 0, 0)
                    curr_date = datetime(*dt_tuple)
                    # print(curr_date)
                    self.dict_per_blueprint[str(init_date)].append(str(curr_date))
                    self.dict_per_blueprint[str(init_date)].extend(
                        [
                            "TOTAL SALES", "CASH SALE", "CREDIT CARD SALES",
                            "PYTM SALE", "LOYALTY REDEEM", "RETURN SALES",
                            "HOLD BILLS", "UNHOLD BILLS", "EXPENSES",
                            "DEEPAWALI PURCHASE", "DRY FRUIT PURCHASE", "STORE RENT",
                            "SUB TOTALS"
                        ]
                    )
                    self.dictionary_updates()
                for cell in range(1, 14 * self.days_per_month_2sqr[i]):
                    if cell <= self.days_per_month_2sqr[i] + 1:
                        self.dict_per_blueprint["BILL HOLDING COLUMN"].append(
                            f"=INDEX($B$8:$B$1000, (ROW(M{cell})-1)*14 + 1)")
                        self.dict_per_blueprint["BILL UNHOLDING COLUMN"].append(
                            f"=INDEX($B$9:$B$1000, (ROW(N{cell})-1)*14 + 1)")
                        self.dict_per_blueprint["TOTALS COLUMN"].append(f"=INDEX($B$2:$B$1000, (ROW(O{cell})-1)*14 + 1)")
                    else:
                        self.dict_per_blueprint["BILL HOLDING COLUMN"].append("")
                        self.dict_per_blueprint["BILL UNHOLDING COLUMN"].append("")
                        self.dict_per_blueprint["TOTALS COLUMN"].append("")
                temp_df = pd.DataFrame(self.dict_per_blueprint)
                self.data_frames[self.monthly_date_chart[i]] = temp_df
        # FIXME: PUT STUFF INTO DICTIONARY ALREADY AND THEN START WITH THE DATE-WISE ITERATIONS (IN BOTH IF-ELSE)
        else:
            for i in range(len(self.days_per_month)):
                # print(" ===== ===== NEXT MONTH ===== =====")
                dt_tuple = (year, i + 1, 1, 0, 0, 0)
                init_date = datetime(*dt_tuple)
                self.dict_per_blueprint = {
                    str(init_date): [
                        "TOTAL SALES", "CASH SALE", "CREDIT CARD SALES",
                        "PYTM SALE", "LOYALTY REDEEM", "RETURN SALES",
                        "HOLD BILLS", "UNHOLD BILLS", "EXPENSES",
                        "DEEPAWALI PURCHASE", "DRY FRUIT PURCHASE", "STORE RENT",
                        "SUB TOTALS"
                    ],
                    "SUMMARY": [
                        0,
                        "=OFFSET(B$2,(ROW()-3)/13*13,0,1,1)-SUM(OFFSET(B$4,(ROW()-3)/13*13,0,10,1))+OFFSET(B$9,(ROW()-3)/13*13,0,1,1)*2",
                        0, 0, 0, 0,
                        "=SUM(OFFSET(F$2, (ROW() - 8)/12*12, 0, 12, 1))",
                        "=SUM(OFFSET(H$2, (ROW() - 9)/12*12, 0, 12, 1))",
                        "=SUM(OFFSET(J$2, (ROW() - 10)/12*12, 0, 12, 1), OFFSET(L$2, (ROW() - 10)/13*13, 0, 13, 1))",
                        0, 0, 0, ""
                    ],
                    "CASH RECEIVED": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "BANK DEPOSIT": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "NAMES OF HOLD BILLS": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        "SUB TOTAL HOLD"
                    ],
                    "AMOUNT HOLD": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        "=SUM(M2:M31)"
                    ],
                    "NAMES OF UNHOLD BILLS": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        "SUB TOTAL UNHOLD"
                    ],
                    "AMOUNT UNHOLD": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        "=SUM(N2:N31)"
                    ],
                    "CASH USED UP BY STORE CONVENIENCES": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        ""
                    ],
                    "AMOUNT USED 1": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "CASH USED UP IN PURCHASE": [
                        "", "", "", "",
                        "", "", "", "",
                        "", "", "", "",
                        ""
                    ],
                    "AMOUNT USED 2": [
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        0, 0, 0, 0,
                        ""
                    ],
                    "BILL HOLDING COLUMN": [],
                    "BILL UNHOLDING COLUMN": [],
                    "TOTALS COLUMN": []
                }

                for days in range(2, self.days_per_month[i] + 1):
                    dt_tuple = (year, i + 1, days, 0, 0, 0)
                    curr_date = datetime(*dt_tuple)
                    # print(curr_date)
                    self.dict_per_blueprint[str(init_date)].append(str(curr_date))
                    self.dict_per_blueprint[str(init_date)].extend(
                        [
                            "TOTAL SALES", "CASH SALE", "CREDIT CARD SALES",
                            "PYTM SALE", "LOYALTY REDEEM", "RETURN SALES",
                            "HOLD BILLS", "UNHOLD BILLS", "EXPENSES",
                            "DEEPAWALI PURCHASE", "DRY FRUIT PURCHASE", "STORE RENT",
                            "SUB TOTALS"
                        ]
                    )
                    self.dictionary_updates()
                for cell in range(1, 14 * self.days_per_month[i]):
                    if cell <= self.days_per_month[i] + 1:
                        self.dict_per_blueprint["BILL HOLDING COLUMN"].append(
                            f"=INDEX($B$8:$B$1000, (ROW(M{cell})-1)*14 + 1)")
                        self.dict_per_blueprint["BILL UNHOLDING COLUMN"].append(
                            f"=INDEX($B$9:$B$1000, (ROW(N{cell})-1)*14 + 1)")
                        self.dict_per_blueprint["TOTALS COLUMN"].append(f"=INDEX($B$2:$B$1000, (ROW(O{cell})-1)*14 + 1)")
                    else:
                        self.dict_per_blueprint["BILL HOLDING COLUMN"].append("")
                        self.dict_per_blueprint["BILL UNHOLDING COLUMN"].append("")
                        self.dict_per_blueprint["TOTALS COLUMN"].append("")

                """
                temp_counter = 0
                for key in self.dict_per_blueprint:
                    temp_counter += 1
                    print(f"{temp_counter}\t", len(self.dict_per_blueprint[key]))
                    if len(self.dict_per_blueprint[key]) == 0:
                        print(key)
                """

                temp_df = pd.DataFrame(self.dict_per_blueprint)
                self.data_frames[self.monthly_date_chart[i]] = temp_df
        # print(" ===== ===== END OF THE YEAR ===== =====")

    def dictionary_updates(self):
        self.dict_per_blueprint["SUMMARY"].append("SUMMARY")
        self.dict_per_blueprint["SUMMARY"].extend(
            [
                0,
                "=OFFSET(B$2,(ROW()-3)/13*13,0,1,1)-SUM(OFFSET(B$4,(ROW()-3)/13*13,0,10,1))+OFFSET(B$9,(ROW()-3)/13*13,0,1,1)*2",
                0, 0, 0, 0,
                "=SUM(OFFSET(F$2, (ROW() - 8)/12*12, 0, 12, 1))",
                "=SUM(OFFSET(H$2, (ROW() - 9)/12*12, 0, 12, 1))",
                "=SUM(OFFSET(J$2, (ROW() - 10)/12*12, 0, 12, 1), OFFSET(L$2, (ROW() - 10)/13*13, 0, 13, 1))",
                0, 0, 0, ""
            ]
        )
        self.dict_per_blueprint["CASH RECEIVED"].append("CASH RECEIVED")
        self.dict_per_blueprint["CASH RECEIVED"].extend(
            [
                0, 0, 0, 0,
                0, 0, 0, 0,
                0, 0, 0, 0,
                ""
            ]
        )
        self.dict_per_blueprint["BANK DEPOSIT"].append("BANK DEPOSIT")
        self.dict_per_blueprint["BANK DEPOSIT"].extend(
            [
                0, 0, 0, 0,
                0, 0, 0, 0,
                0, 0, 0, 0,
                ""
            ]
        )
        self.dict_per_blueprint["NAMES OF HOLD BILLS"].append("NAMES OF HOLD BILLS")
        self.dict_per_blueprint["NAMES OF HOLD BILLS"].extend(
            [
                "", "", "", "",
                "", "", "", "",
                "", "", "", "",
                "SUB TOTAL HOLD"
            ]
        )
        self.dict_per_blueprint["AMOUNT HOLD"].append("AMOUNT HOLD")
        self.dict_per_blueprint["AMOUNT HOLD"].extend(
            [
                0, 0, 0, 0,
                0, 0, 0, 0,
                0, 0, 0, 0,
                "=SUM(M2:M31)"
            ]
        )
        self.dict_per_blueprint["NAMES OF UNHOLD BILLS"].append("NAMES OF UNHOLD BILLS")
        self.dict_per_blueprint["NAMES OF UNHOLD BILLS"].extend(
            [
                "", "", "", "",
                "", "", "", "",
                "", "", "", "",
                "SUB TOTAL UNHOLD"
            ]
        )
        self.dict_per_blueprint["AMOUNT UNHOLD"].append("AMOUNT UNHOLD")
        self.dict_per_blueprint["AMOUNT UNHOLD"].extend(
            [
                0, 0, 0, 0,
                0, 0, 0, 0,
                0, 0, 0, 0,
                "=SUM(N2:N31)"
            ]
        )
        self.dict_per_blueprint["CASH USED UP BY STORE CONVENIENCES"].append(
            "CASH USED UP BY STORE CONVENIENCES")
        self.dict_per_blueprint["CASH USED UP BY STORE CONVENIENCES"].extend(
            [
                "", "", "", "",
                "", "", "", "",
                "", "", "", "",
                ""
            ]
        )
        self.dict_per_blueprint["AMOUNT USED 1"].append("AMOUNT USED 1")
        self.dict_per_blueprint["AMOUNT USED 1"].extend(
            [
                0, 0, 0, 0,
                0, 0, 0, 0,
                0, 0, 0, 0,
                ""
            ]
        )
        self.dict_per_blueprint["CASH USED UP IN PURCHASE"].append("CASH USED UP IN PURCHASE")
        self.dict_per_blueprint["CASH USED UP IN PURCHASE"].extend(
            [
                "", "", "", "",
                "", "", "", "",
                "", "", "", "",
                ""
            ]
        )
        self.dict_per_blueprint["AMOUNT USED 2"].append("AMOUNT USED 2")
        self.dict_per_blueprint["AMOUNT USED 2"].extend(
            [
                0, 0, 0, 0,
                0, 0, 0, 0,
                0, 0, 0, 0,
                ""
            ]
        )

    def func3(self):
        pass

    def component2(self):
        print("*\t\t\tCOMPONENT 2 WORKING...")
        """
        for i in self.data_frames:
            print(self.data_frames[i])
        """

        if os.path.exists(self.filename):
            print(f"* [FAILURE]\tThe file '{self.filename}' already exists,"
                  f" remove/move it elsewhere from the directory to continue..\n\n")

        else:
            print(f"* [SUCCESS]\tThe file '{self.filename}' does not exist, continuing...\n\n")
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                for val in self.monthly_date_chart:
                    # print(f"---- ---- ---- CURRENT MONTH -> {val} ---- ---- ----")
                    self.data_frames[val].to_excel(writer, sheet_name=val, index=False)

            # Load the Excel file using openpyxl
            wb = load_workbook(self.filename)

            # Define styles for the cells (using A1-style referencing)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Blue font color
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green fill
            sky_blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            magenta_fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
            grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            skin_color_fill = PatternFill(start_color="FFDBAC", end_color="FFDBAC", fill_type="solid")

            White_font = Font(color="FFFFFF")

            border_style = Border(
                left=Side(border_style="thin", color="000000"),  # Thin border on the left
                right=Side(border_style="thin", color="000000"),  # Thin border on the right
                top=Side(border_style="thin", color="000000"),  # Thin border on the top
                bottom=Side(border_style="thin", color="000000")  # Thin border on the bottom
            )

            bold_font = Font(bold=True)

            # Apply formatting using A1-style referencing
            temp_counter_2 = 1
            for val in self.monthly_date_chart:
                ws = wb[val]
                dt_tuple = (year, temp_counter_2, 1, 0, 0, 0)
                init_date = datetime(*dt_tuple)
                # print(init_date)
                for column in range(1, len(self.data_frames[val][str(init_date)]) + 2):
                    ws[f'A{column}'].fill = yellow_fill
                    ws[f'B{column}'].fill = grey_fill; ws[f'B{column}'].font = White_font
                    ws[f'C{column}'].fill = sky_blue_fill
                    ws[f'D{column}'].fill = sky_blue_fill
                    ws[f'E{column}'].fill = green_fill
                    ws[f'F{column}'].fill = green_fill
                    ws[f'G{column}'].fill = skin_color_fill
                    ws[f'H{column}'].fill = skin_color_fill
                    ws[f'I{column}'].fill = magenta_fill; ws[f'I{column}'].font = White_font
                    ws[f'J{column}'].fill = magenta_fill; ws[f'J{column}'].font = White_font
                    ws[f'K{column}'].fill = magenta_fill; ws[f'K{column}'].font = White_font
                    ws[f'L{column}'].fill = magenta_fill; ws[f'L{column}'].font = White_font
                    ws[f'M{column}'].fill = grey_fill; ws[f'M{column}'].font = White_font
                    ws[f'N{column}'].fill = grey_fill; ws[f'N{column}'].font = White_font
                    ws[f'O{column}'].fill = grey_fill; ws[f'O{column}'].font = White_font
                    if column % 14 == 1:
                        ws[f'A{column}'].font = bold_font
                        ws[f'B{column}'].font = bold_font
                        ws[f'C{column}'].font = bold_font
                        ws[f'D{column}'].font = bold_font
                        ws[f'E{column}'].font = bold_font
                        ws[f'F{column}'].font = bold_font
                        ws[f'G{column}'].font = bold_font
                        ws[f'H{column}'].font = bold_font
                        ws[f'I{column}'].font = bold_font
                        ws[f'J{column}'].font = bold_font
                        ws[f'K{column}'].font = bold_font
                        ws[f'L{column}'].font = bold_font
                        # ws[f'M{column}'].font = bold_font
                        # ws[f'N{column}'].font = bold_font
                        # ws[f'O{column}'].font = bold_font
                temp_counter_2 += 1
            wb.save(self.filename)
            print("* SAVED SUCCESSFULLY!")

    def main_func(self):
        self.component1()
        self.component2()


year = 2024
filename = f"DATEWISE_SALES_VALIDATED_{year}.xlsx"

obj_val = DSG(filename, year)
obj_val.main_func()
