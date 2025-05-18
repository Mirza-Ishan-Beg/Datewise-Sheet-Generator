import json
import pandas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date, timedelta


class MiniBackend:
    def __init__(self, filename: str, current_year: int, excel_sheet_name: str) -> None:
        self.filename = filename
        self.month_size = []; self.column_size = None
        self.data = self.json_read()
        dates = self.date_iterations(current_year=current_year)
        df_collection = self.excel_dataframe_constructor(self.data, dates)
        self.excel_operator(df_collection, excel_sheet_name)
        self.openpyxl_operator(excel_sheet_name)

    def excel_operator(self, year_data: list[list[list[str]]], file_name: str) -> None:
        """
        Responsible of making the 12 sheets within the excel file.
        """
        all_months = {
            "Jan": [], "Feb": [], "Mar": [], "Apr": [], "May": [], "Jun": [], 
            "Jul": [], "Aug": [], "Sep": [], "Oct": [], "Nov": [], "Dec": [], 
                      }
        months_list = [m for m in all_months.keys()]
        ptr = 0
        for month_data in year_data:
            dataframe = pandas.DataFrame(month_data)
            all_months[months_list[ptr]] = dataframe
            ptr += 1

        # print(all_months)

        with pandas.ExcelWriter(file_name, engine="openpyxl") as pdwriter:
            for month in all_months:
                self.month_size.append(all_months[month].shape[0])
                all_months[month].to_excel(pdwriter, sheet_name=month, index=False, header=False)
        
        return None

    def openpyxl_operator(self, filename: str) -> None:
        """
        After the excel operator executes, it will proceed to open the excel file and decorate it.
        """
        wb = load_workbook(filename)

        ptr_month = 0
        last_col_letter = get_column_letter(self.column_size)
        skip_rows = int(self.month_size[0] / 31)

        for ws in wb.worksheets:
            for date_row in range(1, self.month_size[ptr_month], skip_rows):
                print(f"skip_rows: {date_row}")
                ws.merge_cells(f"A{date_row}:{last_col_letter}{date_row}")
            ptr_month += 1
        
        wb.save(filename)

        return None

    # -----------------------------------------------------------------------------------------------
    # -------------                              UTILITIES                              -------------
    # -----------------------------------------------------------------------------------------------

    def excel_dataframe_constructor(self, data: dict[str, list], dates: list[dict[str, list]]) -> list[list[list[str]]]:
        """
        Responsible for converting dictionary data into Dataframe suitable for pandas.
        data and dates dictionaries are integrated together, then converted to a DF.
        """
        keys = [key for key in data]
        new_data = []
        temp_table_num = len(data[keys[0]]) + 1
        for months in dates:
            for month in months:
                print("================================= Month =================================")
                temp_list = []
                ptr = 0
                for date in months[month]:
                    temp_list.append([date.strftime("%Y-%m-%d")])
                    temp_list[ptr].extend(["" for i in range(len(data[keys[0]])-1)])
                    for row in data[keys[1]]:
                        temp_list.append(row)
                    # print(temp_list)
                    # print(date)
                    ptr += temp_table_num
                new_data.append(temp_list)

        # Structure: 
        # new_data[
        #   temp_list[
        #       row[
        #           string_format_data
        #           ]
        #    ]
        # ]
        self.column_size = len(new_data[0][0])
        return new_data

    def json_read(self) -> dict:
        with open(self.filename, "r") as fr:
            data = json.load(fr)
        return data

    def date_iterations(self, current_year: int) -> list:
        """
        Creates the column for the dates, which will be automatically integrated as the first column.
        This will be a list of them, so that iteration over months is easier when constructing the 
        integrated dataframe(s).
        """
        start_date = date(current_year, 1, 1)
        end_date = date(current_year, 12, 31)
        delta = timedelta(days=1)

        list_of_col = []
        month_ptr = 1
        date_col = {"Date": []}

        while start_date <= end_date:
            date_col["Date"].append(start_date)
            start_date += delta
            if start_date.month > month_ptr:
                list_of_col.append(date_col)
                date_col = {"Date": []}
                month_ptr += 1
        list_of_col.append(date_col)

        return list_of_col
    
    def month_lenght_calculation(self, data: dict[str, list]) -> list[int]:
        """
        Produces the calculated number of rows, used by the openpyxl_operator method to check
        maximum bound of every sheet.
        """
        output = []
        keys = [key for key in data]
        for key in keys:
            temp_table_num = len(data[keys[0]]) + 1
            output.append(temp_table_num)


# obj_minibackend = MiniBackend("save1_alpha.json", 2024, "Temp_sheet_gen.xlsx")
